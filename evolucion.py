import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import re
from datetime import date
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mtick

DATA_DIR    = '/Users/antoniotorres/Desktop/Cosechas_Claude/datos/'
REPORTES    = '/Users/antoniotorres/Desktop/Cosechas_Claude/reportes/'
GRAFICAS    = '/Users/antoniotorres/Desktop/Cosechas_Claude/graficas/'
BASE_FILE   = 'Cosechas1Jun26.xlsx'

MESES = {'Ene':1,'Feb':2,'Mar':3,'Abr':4,'May':5,'Jun':6,
         'Jul':7,'Ago':8,'Sep':9,'Oct':10,'Nov':11,'Dic':12}

# ── helpers ──────────────────────────────────────────────────────────────────

def parse_fecha(fname):
    name = re.sub(r'(CH|MC)$', '', fname.replace('.xlsx','').replace('Cosechas',''))
    m = re.match(r'^(\d+)([A-Za-z]+)(\d+)$', name)
    if not m:
        return None
    day, mes_str, yr = int(m.group(1)), m.group(2), 2000 + int(m.group(3))
    for k, v in MESES.items():
        if mes_str.lower() == k.lower():
            return date(yr, v, day)
    return None

def get_cols(header):
    col = {}
    for i, h in enumerate(header):
        if h in ('Clave Asesor', 'Clave_Asesor_Actual'):  col['id']       = i
        elif h in ('Asesor', 'Asesor_Actual'):            col['nombre']   = i
        elif h in ('Regional', 'RegionalComercial'):      col['regional'] = i
        elif h == 'Sucursal':                             col['sucursal'] = i
        elif h == 'Cosecha':                              col['cosecha']  = i
        elif h in ('Estatus', 'Estatus_Asesor_Actual'):   col['estatus']  = i
    return col

def load_cosechas(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb['Cosechas']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    col = get_cols(rows[0])
    data = {}
    for row in rows[1:]:
        if not any(v is not None for v in row):
            continue
        clave = row[col['id']]
        if clave is None:
            continue
        clave = str(int(float(str(clave)))) if str(clave).replace('.','').isdigit() else str(clave).strip()
        data[clave] = {
            'nombre':   row[col['nombre']],
            'regional': row[col['regional']],
            'sucursal': row[col['sucursal']],
            'cosecha':  row[col['cosecha']],
            'estatus':  row[col.get('estatus', 0)] if 'estatus' in col else None,
        }
    return data

def col_label(d):
    meses_es = {1:'Ene',2:'Feb',3:'Mar',4:'Abr',5:'May',6:'Jun',
                7:'Jul',8:'Ago',9:'Sep',10:'Oct',11:'Nov',12:'Dic'}
    return f"{d.day:02d}-{meses_es[d.month]}-{str(d.year)[2:]}"

# ── ordenar archivos por fecha ────────────────────────────────────────────────

all_files = [f for f in os.listdir(DATA_DIR) if f.endswith('.xlsx') and not f.startswith('~$')]

dated = {}
for f in all_files:
    d = parse_fecha(f)
    if d is None:
        continue
    is_mc = 'MC' in f.replace('Cosechas','').replace('.xlsx','')
    if d not in dated:
        dated[d] = (f, is_mc)
    else:
        # prefer non-MC
        if dated[d][1] and not is_mc:
            dated[d] = (f, is_mc)

ordered = sorted(dated.keys())
print(f"Semanas encontradas: {len(ordered)}")
for d in ordered:
    print(f"  {col_label(d):12s} → {dated[d][0]}")

# ── cargar todos los datos ────────────────────────────────────────────────────

all_data = {}   # date -> {clave -> {...}}
for d in ordered:
    fname = dated[d][0]
    all_data[d] = load_cosechas(os.path.join(DATA_DIR, fname))
    print(f"Cargado {fname}: {len(all_data[d])} asesores")

# ── asesores activos del archivo base ────────────────────────────────────────

base_date = date(2026, 6, 1)
base_data = all_data[base_date]
activos = {k: v for k, v in base_data.items() if str(v.get('estatus','')).lower() == 'activo'}
print(f"\nAsesores activos en base: {len(activos)}")

# ── construir tabla de evolución ──────────────────────────────────────────────

# Columnas de semanas (todas excepto la base — la base va al final)
semanas = ordered  # base_date ya está incluida al final

rows_tabla = []
for clave, info in sorted(activos.items(), key=lambda x: (str(x[1]['regional']), str(x[1]['sucursal']), str(x[1]['nombre']))):
    row = {
        'regional': info['regional'],
        'sucursal': info['sucursal'],
        'clave':    clave,
        'nombre':   info['nombre'],
    }
    for d in semanas:
        asesor_en_semana = all_data[d].get(clave)
        row[col_label(d)] = asesor_en_semana['cosecha'] if asesor_en_semana else None
    rows_tabla.append(row)

print(f"Filas en tabla de evolución: {len(rows_tabla)}")

# ── estilos ───────────────────────────────────────────────────────────────────

HDR_FILL   = PatternFill('solid', fgColor='1F4E79')
HDR_FONT   = Font(color='FFFFFF', bold=True, size=10)
FIXED_FILL = PatternFill('solid', fgColor='D6E4F0')
FIXED_FONT = Font(bold=True, size=10)
UP_FILL    = PatternFill('solid', fgColor='C6EFCE')
DN_FILL    = PatternFill('solid', fgColor='FFC7CE')
EQ_FILL    = PatternFill('solid', fgColor='FFEB9C')
THIN       = Border(
    left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'))

def pct(v):
    if v is None: return ''
    try:    return f"{float(v)*100:.1f}%"
    except: return str(v)

# ── Excel ─────────────────────────────────────────────────────────────────────

wb_out = openpyxl.Workbook()

# ── Hoja 1: Evolución semana a semana ─────────────────────────────────────────

ws_evo = wb_out.active
ws_evo.title = 'Evolución'

semana_labels = [col_label(d) for d in semanas]
headers = ['Regional', 'Sucursal', 'Clave Asesor', 'Nombre Asesor'] + semana_labels

# encabezados
for ci, h in enumerate(headers, 1):
    cell = ws_evo.cell(row=1, column=ci, value=h)
    cell.font      = HDR_FONT
    cell.fill      = HDR_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border    = THIN

# filas de datos
for ri, row in enumerate(rows_tabla, 2):
    fixed_vals = [row['regional'], row['sucursal'], row['clave'], row['nombre']]
    for ci, v in enumerate(fixed_vals, 1):
        c = ws_evo.cell(row=ri, column=ci, value=v)
        c.fill = FIXED_FILL; c.font = FIXED_FONT; c.border = THIN
        c.alignment = Alignment(horizontal='left')

    semana_vals = [row.get(lbl) for lbl in semana_labels]
    for ci, v in enumerate(semana_vals, 5):
        c = ws_evo.cell(row=ri, column=ci, value=v)
        c.border = THIN
        c.alignment = Alignment(horizontal='center')
        if v is not None:
            try:
                c.number_format = '0.00%'
                c.value = float(v)
            except:
                pass

# anchos de columna
ws_evo.column_dimensions['A'].width = 18
ws_evo.column_dimensions['B'].width = 18
ws_evo.column_dimensions['C'].width = 12
ws_evo.column_dimensions['D'].width = 32
for ci in range(5, len(headers) + 1):
    ws_evo.column_dimensions[get_column_letter(ci)].width = 11

ws_evo.freeze_panes = 'E2'

# ── Hoja 2: Resumen ──────────────────────────────────────────────────────────

ws_res = wb_out.create_sheet('Resumen')

res_headers = ['Regional', 'Sucursal', 'Clave Asesor', 'Nombre Asesor',
               'Cosecha Inicial', 'Cosecha Final', 'Variación', 'Tendencia']
for ci, h in enumerate(res_headers, 1):
    cell = ws_res.cell(row=1, column=ci, value=h)
    cell.font = HDR_FONT; cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = THIN

res_rows = []
for row in rows_tabla:
    vals = [row.get(lbl) for lbl in semana_labels]
    nums = [(i, float(v)) for i, v in enumerate(vals) if v is not None]
    if len(nums) < 2:
        continue
    ini_idx, ini = nums[0]
    fin_idx, fin = nums[-1]
    variacion = fin - ini
    if variacion > 0.005:
        tendencia = 'Subió'
        fill = UP_FILL
    elif variacion < -0.005:
        tendencia = 'Bajó'
        fill = DN_FILL
    else:
        tendencia = 'Se mantuvo'
        fill = EQ_FILL
    res_rows.append((row['regional'], row['sucursal'], row['clave'], row['nombre'],
                     ini, fin, variacion, tendencia, fill))

# ordenar por tendencia luego por variación desc
order = {'Bajó': 0, 'Se mantuvo': 1, 'Subió': 2}
res_rows.sort(key=lambda x: (order[x[7]], -x[6]))

for ri, r in enumerate(res_rows, 2):
    fill = r[8]
    vals = r[:8]
    for ci, v in enumerate(vals, 1):
        c = ws_res.cell(row=ri, column=ci, value=v)
        c.border = THIN
        c.alignment = Alignment(horizontal='center' if ci > 4 else 'left')
        if ci in (5, 6, 7):
            try:
                c.number_format = '0.00%'
                c.value = float(v)
            except:
                pass
        if ci >= 5:
            c.fill = fill

# anchos
for ci, w in zip(range(1, 9), [18, 18, 12, 32, 14, 14, 12, 14]):
    ws_res.column_dimensions[get_column_letter(ci)].width = w

# conteo resumen al final
totales = {'Subió': 0, 'Bajó': 0, 'Se mantuvo': 0}
for r in res_rows:
    totales[r[7]] += 1

blank_row = len(res_rows) + 3
ws_res.cell(row=blank_row,   column=1, value='RESUMEN').font = Font(bold=True)
ws_res.cell(row=blank_row+1, column=1, value='Subió').fill    = UP_FILL
ws_res.cell(row=blank_row+1, column=2, value=totales['Subió'])
ws_res.cell(row=blank_row+2, column=1, value='Bajó').fill     = DN_FILL
ws_res.cell(row=blank_row+2, column=2, value=totales['Bajó'])
ws_res.cell(row=blank_row+3, column=1, value='Se mantuvo').fill = EQ_FILL
ws_res.cell(row=blank_row+3, column=2, value=totales['Se mantuvo'])

# ── guardar Excel ─────────────────────────────────────────────────────────────

out_path = os.path.join(REPORTES, 'Evolucion_Asesores.xlsx')
wb_out.save(out_path)
print(f"\nExcel guardado: {out_path}")

# ── gráficas ──────────────────────────────────────────────────────────────────

plt.style.use('seaborn-v0_8-whitegrid')
COLS_GRAFICAS = 4
graficas_creadas = 0

for row in rows_tabla:
    vals = [(semana_labels[i], float(v)) for i, v in enumerate([row.get(lbl) for lbl in semana_labels]) if v is not None]
    if len(vals) < 2:
        continue

    labels_x = [v[0] for v in vals]
    cosechas  = [v[1] for v in vals]

    fig, ax = plt.subplots(figsize=(10, 4))
    ax.plot(labels_x, cosechas, marker='o', linewidth=2, markersize=5, color='#1F4E79')
    ax.fill_between(range(len(cosechas)), cosechas, alpha=0.08, color='#1F4E79')

    # línea de referencia 90%
    ax.axhline(0.9, color='#C00000', linestyle='--', linewidth=1, label='Meta 90%')

    ax.yaxis.set_major_formatter(mtick.PercentFormatter(xmax=1.0, decimals=1))
    ax.set_xticks(range(len(labels_x)))
    ax.set_xticklabels(labels_x, rotation=45, ha='right', fontsize=8)
    ax.set_title(f"{row['nombre']}  |  Clave {row['clave']}  |  {row['sucursal']}", fontsize=10, pad=8)
    ax.set_ylabel('Cosecha', fontsize=9)
    ax.legend(fontsize=8)
    plt.tight_layout()

    safe_name = re.sub(r'[^\w\-]', '_', str(row['nombre']))
    path = os.path.join(GRAFICAS, f"{row['clave']}_{safe_name}.png")
    fig.savefig(path, dpi=120, bbox_inches='tight')
    plt.close(fig)
    graficas_creadas += 1

print(f"Gráficas creadas: {graficas_creadas}")
print("\nResumen tendencias:")
for k, v in totales.items():
    print(f"  {k}: {v}")
print("\nListo.")
