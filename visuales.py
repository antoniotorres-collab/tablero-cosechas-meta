import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
import os, re, tempfile
from datetime import date
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mtick
import matplotlib.colors as mcolors
import numpy as np

DATA_DIR  = '/Users/antoniotorres/Desktop/Cosechas_Claude/datos/'
REPORTES  = '/Users/antoniotorres/Desktop/Cosechas_Claude/reportes/'
GRAFICAS  = '/Users/antoniotorres/Desktop/Cosechas_Claude/graficas/'
EXCEL_OUT = os.path.join(REPORTES, 'Evolucion_Asesores.xlsx')

MESES = {'Ene':1,'Feb':2,'Mar':3,'Abr':4,'May':5,'Jun':6,
         'Jul':7,'Ago':8,'Sep':9,'Oct':10,'Nov':11,'Dic':12}

# ── helpers (mismo que evolucion.py) ─────────────────────────────────────────

def parse_fecha(fname):
    name = re.sub(r'(CH|MC)$', '', fname.replace('.xlsx','').replace('Cosechas',''))
    m = re.match(r'^(\d+)([A-Za-z]+)(\d+)$', name)
    if not m: return None
    day, mes_str, yr = int(m.group(1)), m.group(2), 2000 + int(m.group(3))
    for k, v in MESES.items():
        if mes_str.lower() == k.lower(): return date(yr, v, day)
    return None

def get_cols(header):
    col = {}
    for i, h in enumerate(header):
        if h in ('Clave Asesor','Clave_Asesor_Actual'):  col['id']       = i
        elif h in ('Asesor','Asesor_Actual'):            col['nombre']   = i
        elif h in ('Regional','RegionalComercial'):      col['regional'] = i
        elif h == 'Sucursal':                            col['sucursal'] = i
        elif h == 'Cosecha':                             col['cosecha']  = i
        elif h in ('Estatus','Estatus_Asesor_Actual'):   col['estatus']  = i
    return col

def load_cosechas(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb['Cosechas']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    col = get_cols(rows[0])
    data = {}
    for row in rows[1:]:
        if not any(v is not None for v in row): continue
        clave = row[col['id']]
        if clave is None: continue
        clave = str(int(float(str(clave)))) if str(clave).replace('.','').isdigit() else str(clave).strip()
        data[clave] = {
            'nombre':   row[col['nombre']],
            'regional': row[col['regional']],
            'sucursal': row[col['sucursal']],
            'cosecha':  row[col['cosecha']],
            'estatus':  row[col['estatus']] if 'estatus' in col else None,
        }
    return data

def col_label(d):
    meses_es = {1:'Ene',2:'Feb',3:'Mar',4:'Abr',5:'May',6:'Jun',
                7:'Jul',8:'Ago',9:'Sep',10:'Oct',11:'Nov',12:'Dic'}
    return f"{d.day:02d}-{meses_es[d.month]}-{str(d.year)[2:]}"

# ── cargar datos ──────────────────────────────────────────────────────────────

all_files = [f for f in os.listdir(DATA_DIR) if f.endswith('.xlsx') and not f.startswith('~$')]
dated = {}
for f in all_files:
    d = parse_fecha(f)
    if d is None: continue
    is_mc = 'MC' in f.replace('Cosechas','').replace('.xlsx','')
    if d not in dated:
        dated[d] = (f, is_mc)
    elif dated[d][1] and not is_mc:
        dated[d] = (f, is_mc)

ordered = sorted(dated.keys())
semana_labels = [col_label(d) for d in ordered]

all_data = {}
for d in ordered:
    all_data[d] = load_cosechas(os.path.join(DATA_DIR, dated[d][0]))

base_date = date(2026, 6, 1)
base_data = all_data[base_date]
activos   = {k: v for k, v in base_data.items() if str(v.get('estatus','')).lower() == 'activo'}

# tabla de evolución
rows_tabla = []
for clave, info in sorted(activos.items(), key=lambda x: (str(x[1]['regional']), str(x[1]['nombre']))):
    row = {'regional': info['regional'], 'sucursal': info['sucursal'],
           'clave': clave, 'nombre': info['nombre']}
    for d in ordered:
        ae = all_data[d].get(clave)
        row[col_label(d)] = float(ae['cosecha']) if ae and ae['cosecha'] is not None else None
    rows_tabla.append(row)

print(f"Asesores activos: {len(rows_tabla)} | Semanas: {len(ordered)}")

# ── paleta y estilo ───────────────────────────────────────────────────────────

AZUL   = '#1F4E79'
VERDE  = '#375623'
ROJO   = '#C00000'
NARANJA= '#E26B0A'
GRIS   = '#595959'
plt.rcParams.update({'font.family': 'DejaVu Sans', 'axes.spines.top': False,
                     'axes.spines.right': False})

tmp_files = []

def save_tmp(fig, name):
    path = os.path.join(GRAFICAS, name)
    fig.savefig(path, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    tmp_files.append(path)
    print(f"  Guardada: {name}")
    return path

# ══════════════════════════════════════════════════════════════════════════════
# GRÁFICA 1 — Dashboard resumen (2×2)
# ══════════════════════════════════════════════════════════════════════════════

fig, axes = plt.subplots(2, 2, figsize=(16, 10))
fig.suptitle('Dashboard Cosechas — Portafolio Activo  |  Ene–Jun 2026',
             fontsize=14, fontweight='bold', color=AZUL, y=1.01)

# ── 1a: promedio semanal del portafolio ──────────────────────────────────────
ax = axes[0, 0]
promedios = []
for lbl in semana_labels:
    vals = [r[lbl] for r in rows_tabla if r[lbl] is not None]
    promedios.append(np.mean(vals) if vals else None)

xs = range(len(semana_labels))
ys = [v if v else np.nan for v in promedios]
ax.plot(xs, ys, marker='o', color=AZUL, linewidth=2, markersize=4)
ax.fill_between(xs, ys, alpha=0.1, color=AZUL)
ax.axhline(0.9, color=ROJO, linestyle='--', linewidth=1.2, label='Meta 90%')
ax.set_xticks(list(xs))
ax.set_xticklabels(semana_labels, rotation=45, ha='right', fontsize=7)
ax.yaxis.set_major_formatter(mtick.PercentFormatter(xmax=1.0, decimals=1))
ax.set_title('Promedio semanal del portafolio', fontweight='bold', fontsize=10, color=GRIS)
ax.legend(fontsize=8)
# anotar el valor de cada semana
for x, y in zip(xs, ys):
    if not np.isnan(y):
        ax.annotate(f'{y*100:.1f}%', (x, y), textcoords='offset points',
                    xytext=(0, 5), fontsize=6, ha='center', color=AZUL)

# ── 1b: donut de tendencias ──────────────────────────────────────────────────
ax = axes[0, 1]
totales = {'Subió': 0, 'Se mantuvo': 0, 'Bajó': 0}
for row in rows_tabla:
    nums = [row[lbl] for lbl in semana_labels if row[lbl] is not None]
    if len(nums) < 2: continue
    v = nums[-1] - nums[0]
    if v > 0.005:   totales['Subió'] += 1
    elif v < -0.005: totales['Bajó'] += 1
    else:           totales['Se mantuvo'] += 1

labels_d = list(totales.keys())
sizes_d  = list(totales.values())
colors_d = ['#375623', '#FFEB9C', '#C00000']
wedges, texts, autotexts = ax.pie(
    sizes_d, labels=None, autopct='%1.0f%%', startangle=90,
    colors=colors_d, pctdistance=0.75,
    wedgeprops=dict(width=0.5, edgecolor='white', linewidth=2))
for at in autotexts:
    at.set_fontsize(11); at.set_fontweight('bold'); at.set_color('white')
ax.legend(wedges, [f'{l}: {s}' for l, s in zip(labels_d, sizes_d)],
          loc='lower center', bbox_to_anchor=(0.5, -0.08), ncol=3, fontsize=9)
ax.set_title('Tendencia Ene → Jun', fontweight='bold', fontsize=10, color=GRIS)

# ── 1c: promedio por regional — última semana ────────────────────────────────
ax = axes[1, 0]
ultima = semana_labels[-1]
by_reg = {}
for row in rows_tabla:
    reg = str(row['regional'])
    v   = row[ultima]
    if v is not None:
        by_reg.setdefault(reg, []).append(v)
reg_names = sorted(by_reg.keys())
reg_means = [np.mean(by_reg[r]) for r in reg_names]
colors_bar = [VERDE if m >= 0.9 else ROJO for m in reg_means]
bars = ax.barh(reg_names, reg_means, color=colors_bar, edgecolor='white', height=0.6)
ax.axvline(0.9, color=NARANJA, linestyle='--', linewidth=1.2, label='Meta 90%')
ax.xaxis.set_major_formatter(mtick.PercentFormatter(xmax=1.0, decimals=1))
for bar, val in zip(bars, reg_means):
    ax.text(val + 0.002, bar.get_y() + bar.get_height()/2,
            f'{val*100:.1f}%', va='center', fontsize=9, color=GRIS)
ax.set_title(f'Promedio por Regional  ({ultima})', fontweight='bold', fontsize=10, color=GRIS)
ax.legend(fontsize=8)
ax.set_xlim(0, 1.05)

# ── 1d: distribución de cosechas — última semana ─────────────────────────────
ax = axes[1, 1]
vals_dist = [row[ultima] for row in rows_tabla if row[ultima] is not None]
bins = np.arange(0, 1.05, 0.05)
n, bins_out, patches = ax.hist(vals_dist, bins=bins, color=AZUL, edgecolor='white', alpha=0.85)
for patch, left in zip(patches, bins_out[:-1]):
    if left >= 0.9: patch.set_facecolor(VERDE)
    elif left < 0.7: patch.set_facecolor(ROJO)
ax.axvline(0.9, color=ROJO, linestyle='--', linewidth=1.2, label='Meta 90%')
ax.xaxis.set_major_formatter(mtick.PercentFormatter(xmax=1.0, decimals=0))
ax.set_xlabel('Cosecha', fontsize=9); ax.set_ylabel('# Asesores', fontsize=9)
ax.set_title(f'Distribución de Cosechas  ({ultima})', fontweight='bold', fontsize=10, color=GRIS)
ax.legend(fontsize=8)

plt.tight_layout()
path_dashboard = save_tmp(fig, 'dashboard_resumen.png')

# ══════════════════════════════════════════════════════════════════════════════
# GRÁFICA 2 — Heatmap asesor × semana
# ══════════════════════════════════════════════════════════════════════════════

n_asesores = len(rows_tabla)
n_semanas  = len(semana_labels)
matrix = np.full((n_asesores, n_semanas), np.nan)
for i, row in enumerate(rows_tabla):
    for j, lbl in enumerate(semana_labels):
        v = row[lbl]
        if v is not None:
            matrix[i, j] = v

# etiquetas cortas para el eje Y
nombre_labels = [f"{r['clave']} {r['nombre'][:22]}" for r in rows_tabla]

fig_h = max(12, n_asesores * 0.28)
fig, ax = plt.subplots(figsize=(18, fig_h))

cmap = mcolors.LinearSegmentedColormap.from_list(
    'cosecha', ['#C00000', '#FF9900', '#FFEB9C', '#92D050', '#375623'])
cmap.set_bad(color='#F2F2F2')

im = ax.imshow(matrix, aspect='auto', cmap=cmap, vmin=0.6, vmax=1.0,
               interpolation='nearest')

# línea vertical de referencia en el último corte
ax.axvline(n_semanas - 1.5, color='white', linewidth=2, linestyle='--', alpha=0.6)

ax.set_xticks(range(n_semanas))
ax.set_xticklabels(semana_labels, rotation=45, ha='right', fontsize=8)
ax.set_yticks(range(n_asesores))
ax.set_yticklabels(nombre_labels, fontsize=7)

# valores en cada celda
for i in range(n_asesores):
    for j in range(n_semanas):
        v = matrix[i, j]
        if not np.isnan(v):
            color_txt = 'white' if v < 0.75 or v > 0.97 else '#1a1a1a'
            ax.text(j, i, f'{v*100:.0f}', ha='center', va='center',
                    fontsize=5.5, color=color_txt)

cbar = fig.colorbar(im, ax=ax, fraction=0.015, pad=0.01)
cbar.ax.yaxis.set_major_formatter(mtick.PercentFormatter(xmax=1.0, decimals=0))
cbar.set_label('Cosecha', fontsize=9)

ax.set_title('Heatmap de Cosecha por Asesor y Semana  |  Ene–Jun 2026',
             fontweight='bold', fontsize=12, color=AZUL, pad=12)
plt.tight_layout()
path_heatmap = save_tmp(fig, 'heatmap_asesores.png')

# ══════════════════════════════════════════════════════════════════════════════
# GRÁFICA 3 — Top 10 mejores y peores
# ══════════════════════════════════════════════════════════════════════════════

variaciones = []
for row in rows_tabla:
    nums = [(lbl, row[lbl]) for lbl in semana_labels if row[lbl] is not None]
    if len(nums) < 2: continue
    var = nums[-1][1] - nums[0][1]
    variaciones.append({'nombre': row['nombre'], 'clave': row['clave'],
                        'sucursal': row['sucursal'], 'var': var,
                        'ini': nums[0][1], 'fin': nums[-1][1]})

variaciones.sort(key=lambda x: x['var'])
peores = variaciones[:10]
mejores = variaciones[-10:][::-1]

fig, (ax_l, ax_r) = plt.subplots(1, 2, figsize=(16, 7))
fig.suptitle('Top 10 Mejores y Peores Asesores  |  Variación Ene → Jun 2026',
             fontsize=13, fontweight='bold', color=AZUL)

def short_label(r):
    parts = r['nombre'].split()
    return f"{parts[0]} {parts[-1]}  ({r['clave']})"

# peores (izquierda)
names_p = [short_label(r) for r in peores]
vars_p  = [r['var'] for r in peores]
bars_p  = ax_l.barh(names_p, vars_p, color=ROJO, edgecolor='white', height=0.6)
ax_l.axvline(0, color=GRIS, linewidth=0.8)
ax_l.xaxis.set_major_formatter(mtick.PercentFormatter(xmax=1.0, decimals=1))
for bar, r in zip(bars_p, peores):
    ax_l.text(r['var'] - 0.001, bar.get_y() + bar.get_height()/2,
              f"{r['var']*100:.1f}%", va='center', ha='right', fontsize=8.5,
              color='white', fontweight='bold')
ax_l.set_title('⬇ Los que más bajaron', fontsize=10, color=ROJO, fontweight='bold')
ax_l.invert_yaxis()

# mejores (derecha)
names_m = [short_label(r) for r in mejores]
vars_m  = [r['var'] for r in mejores]
bars_m  = ax_r.barh(names_m, vars_m, color=VERDE, edgecolor='white', height=0.6)
ax_r.axvline(0, color=GRIS, linewidth=0.8)
ax_r.xaxis.set_major_formatter(mtick.PercentFormatter(xmax=1.0, decimals=1))
for bar, r in zip(bars_m, mejores):
    ax_r.text(r['var'] + 0.001, bar.get_y() + bar.get_height()/2,
              f"+{r['var']*100:.1f}%", va='center', ha='left', fontsize=8.5,
              color='white', fontweight='bold')
ax_r.set_title('⬆ Los que más subieron', fontsize=10, color=VERDE, fontweight='bold')
ax_r.invert_yaxis()

plt.tight_layout()
path_ranking = save_tmp(fig, 'top_ranking.png')

# ══════════════════════════════════════════════════════════════════════════════
# GRÁFICA 4 — Evolución por regional
# ══════════════════════════════════════════════════════════════════════════════

regionales = sorted(set(str(r['regional']) for r in rows_tabla))
n_reg = len(regionales)
cols_reg = 2
rows_reg = (n_reg + 1) // 2

fig, axes_reg = plt.subplots(rows_reg, cols_reg,
                              figsize=(16, rows_reg * 4), squeeze=False)
fig.suptitle('Evolución de Cosecha por Regional  |  Ene–Jun 2026',
             fontsize=13, fontweight='bold', color=AZUL, y=1.01)

COLOR_POOL = ['#1F4E79','#375623','#E26B0A','#7030A0','#C00000',
              '#2E75B6','#548235','#BF8F00','#833C00']

for idx, reg in enumerate(regionales):
    r_idx, c_idx = divmod(idx, cols_reg)
    ax = axes_reg[r_idx][c_idx]
    asesores_reg = [r for r in rows_tabla if str(r['regional']) == reg]

    # línea de cada asesor (gris, fina)
    for row in asesores_reg:
        ys = [row[lbl] for lbl in semana_labels]
        ax.plot(range(len(semana_labels)), ys, color='#BFBFBF', linewidth=0.7, alpha=0.6)

    # promedio del regional (grueso, coloreado)
    prom_reg = []
    for lbl in semana_labels:
        vs = [row[lbl] for row in asesores_reg if row[lbl] is not None]
        prom_reg.append(np.mean(vs) if vs else np.nan)

    ax.plot(range(len(semana_labels)), prom_reg, color=COLOR_POOL[idx % len(COLOR_POOL)],
            linewidth=2.5, marker='o', markersize=4, label='Promedio regional', zorder=5)
    ax.axhline(0.9, color=ROJO, linestyle='--', linewidth=1, label='Meta 90%')

    ax.set_xticks(range(len(semana_labels)))
    ax.set_xticklabels(semana_labels, rotation=45, ha='right', fontsize=7)
    ax.yaxis.set_major_formatter(mtick.PercentFormatter(xmax=1.0, decimals=1))
    ax.set_ylim(0.5, 1.05)
    ax.set_title(f'{reg}  ({len(asesores_reg)} asesores)',
                 fontweight='bold', fontsize=10, color=GRIS)
    ax.legend(fontsize=7)

# ocultar subplots vacíos
for idx in range(n_reg, rows_reg * cols_reg):
    r_idx, c_idx = divmod(idx, cols_reg)
    axes_reg[r_idx][c_idx].set_visible(False)

plt.tight_layout()
path_regional = save_tmp(fig, 'por_regional.png')

# ══════════════════════════════════════════════════════════════════════════════
# INSERTAR EN EXCEL — hojas de dashboard
# ══════════════════════════════════════════════════════════════════════════════

print("\nAbriendo Excel para insertar gráficas...")
wb = openpyxl.load_workbook(EXCEL_OUT)

# quitar hojas de dashboard previas si existen
for sname in ['Dashboard', 'Heatmap', 'Por Regional', 'Ranking']:
    if sname in wb.sheetnames:
        del wb[sname]

TITLE_FONT = Font(name='Calibri', size=14, bold=True, color='1F4E79')
TITLE_FILL = PatternFill('solid', fgColor='D6E4F0')

def add_chart_sheet(wb, sheet_name, img_path, title):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    # título
    ws['B2'] = title
    ws['B2'].font = TITLE_FONT
    ws['B2'].fill = TITLE_FILL
    ws['B2'].alignment = Alignment(horizontal='left')
    ws.row_dimensions[2].height = 22
    # imagen
    img = XLImage(img_path)
    ws.add_image(img, 'B4')
    return ws

add_chart_sheet(wb, 'Dashboard',   path_dashboard, '📊 Dashboard Resumen — Portafolio Activo')
add_chart_sheet(wb, 'Heatmap',     path_heatmap,   '🌡 Heatmap de Cosecha por Asesor y Semana')
add_chart_sheet(wb, 'Ranking',     path_ranking,   '🏆 Top 10 Mejores y Peores Asesores')
add_chart_sheet(wb, 'Por Regional',path_regional,  '📈 Evolución por Regional')

# reordenar hojas: primero las nuevas, luego los datos
sheet_order = ['Dashboard', 'Heatmap', 'Ranking', 'Por Regional', 'Evolución', 'Resumen']
for name in sheet_order:
    if name in wb.sheetnames:
        wb.move_sheet(name, offset=-len(wb.sheetnames))

wb.save(EXCEL_OUT)
print(f"Excel actualizado: {EXCEL_OUT}")
print("\nHojas en el archivo:")
for s in wb.sheetnames:
    print(f"  • {s}")
